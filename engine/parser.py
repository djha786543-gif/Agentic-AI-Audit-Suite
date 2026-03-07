"""
engine/parser.py
────────────────
Universal Audit Data Parser.

Supports:
  - CSV (any delimiter, auto-detected)
  - Excel (XLSX, XLS) - all sheets
  - JSON / JSONL
  - SAP exported TXT (tab-delimited, ALV export format)
  - SAP IDOC-style flat files
  - Auto-detection of data type (users, changes, transactions, etc.)
"""
from __future__ import annotations
import json
import io
import re
from typing import List, Dict, Any, Tuple, Optional
import logging

logger = logging.getLogger(__name__)

# Column name aliases → normalized field names
COLUMN_ALIASES = {
    # User/Access fields
    "userid": "user_id", "username": "user_id", "user name": "user_id",
    "sap user": "user_id", "employee id": "user_id", "emp id": "user_id",
    "login": "user_id", "logon": "user_id", "account": "user_id",
    "last logon": "last_login_date", "last login": "last_login_date",
    "last logon date": "last_login_date", "last access": "last_login_date",
    "termination date": "termination_date", "term date": "termination_date",
    "separation date": "termination_date", "end date": "termination_date",
    "account status": "status", "user status": "status", "active": "status",
    "role": "roles", "roles": "roles", "profile": "roles", "authorization": "roles",
    "composite role": "roles", "single role": "roles",
    "mfa": "mfa_enabled", "2fa": "mfa_enabled", "multi factor": "mfa_enabled",
    "last review": "access_review_date", "review date": "access_review_date",
    "access review": "access_review_date", "cert date": "access_review_date",

    # Change management fields
    "change number": "ticket_id", "change no": "ticket_id", "chg": "ticket_id",
    "ticket": "ticket_id", "ticket number": "ticket_id", "cr number": "ticket_id",
    "requested by": "initiator", "created by": "initiator", "requester": "initiator",
    "approved by": "approver", "change approver": "approver", "approval by": "approver",
    "implementation date": "implementation_date", "deploy date": "implementation_date",
    "go live date": "implementation_date", "actual start": "implementation_date",
    "change type": "change_type", "type": "change_type",
    "uat complete": "test_evidence", "test evidence": "test_evidence",
    "testing complete": "test_evidence", "tested": "test_evidence",
    "post implementation review": "post_impl_review", "pir": "post_impl_review",

    # Transaction / AP fields
    "invoice number": "invoice_id", "invoice no": "invoice_id", "inv number": "invoice_id",
    "document number": "invoice_id", "doc no": "invoice_id",
    "vendor name": "vendor", "supplier": "vendor", "supplier name": "vendor",
    "invoice amount": "invoice_amount", "gross amount": "invoice_amount",
    "net amount": "invoice_amount", "amount": "invoice_amount",
    "po amount": "po_amount", "purchase order amount": "po_amount",
    "gr amount": "gr_amount", "goods receipt amount": "gr_amount",
    "invoice date": "invoice_date", "posting date": "invoice_date",
    "doc date": "invoice_date",
    "payment date": "payment_date", "due date": "payment_date",
    "three way match": "three_way_match", "3 way match": "three_way_match",
    "match status": "three_way_match",

    # Backup / Operations fields
    "job name": "job_name", "backup job": "job_name",
    "backup date": "backup_date", "completed": "backup_date",
    "restore tested": "restore_tested", "verified": "restore_tested",
    "job status": "status", "backup status": "status", "result": "status",
}


def _normalize_col(col: str) -> str:
    """Normalize column name for alias lookup."""
    return col.strip().lower().replace("_", " ").replace("-", " ")


def _map_columns(row: Dict) -> Dict:
    """Remap column names using the alias table."""
    mapped = {}
    for k, v in row.items():
        norm = _normalize_col(str(k))
        canonical = COLUMN_ALIASES.get(norm, norm.replace(" ", "_"))
        mapped[canonical] = v
    return mapped


def _detect_delimiter(content: str) -> str:
    """Auto-detect CSV delimiter."""
    sample = content[:2000]
    counts = {
        ",": sample.count(","),
        "\t": sample.count("\t"),
        "|": sample.count("|"),
        ";": sample.count(";"),
    }
    return max(counts, key=counts.get)


def _detect_data_type(columns: List[str]) -> str:
    """Guess what kind of audit data this file contains."""
    cols = " ".join(c.lower() for c in columns)

    if any(k in cols for k in ["user_id", "username", "logon", "role", "profile", "mfa"]):
        return "users"
    if any(k in cols for k in ["ticket_id", "change_number", "chg", "approver", "change_type"]):
        return "changes"
    if any(k in cols for k in ["invoice", "vendor", "po_amount", "three_way", "payment"]):
        return "transactions"
    if any(k in cols for k in ["backup", "restore", "job_name", "backup_date"]):
        return "backup"
    if any(k in cols for k in ["incident", "priority", "resolved", "sla"]):
        return "incident"
    if any(k in cols for k in ["interface", "source_count", "target_count", "records_sent"]):
        return "interfaces"
    if any(k in cols for k in ["hr", "employee_name", "employment_status", "legal_entity", "personnel"]):
        return "hr_master"
    if any(k in cols for k in ["dr_test", "rto", "rpo", "disaster"]):
        return "dr"
    return "unknown"


def parse_csv(content: str) -> Tuple[List[Dict], str]:
    """Parse CSV content, return (records, data_type)."""
    import csv
    delimiter = _detect_delimiter(content)
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    records = []
    for idx, row in enumerate(reader, start=2):
        mapped = _map_columns(row)
        # Preserve source-to-report lineage at row granularity.
        mapped["_lineage"] = {
            "source_type": "csv",
            "row_number": idx,
        }
        records.append(mapped)
    data_type = _detect_data_type(reader.fieldnames or [])
    logger.info("CSV parsed: %d records, type=%s, delimiter='%s'", len(records), data_type, repr(delimiter))
    return records, data_type


def parse_excel(file_bytes: bytes) -> Dict[str, Tuple[List[Dict], str]]:
    """Parse Excel file, return dict of {sheet_name: (records, data_type)}."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            records = []
            for row_idx, row in enumerate(rows[1:], start=2):
                if all(v is None for v in row):
                    continue
                raw = dict(zip(headers, row))
                mapped = _map_columns(raw)
                mapped["_lineage"] = {
                    "source_type": "excel",
                    "sheet": sheet_name,
                    "row_number": row_idx,
                }
                records.append(mapped)
            data_type = _detect_data_type(headers)
            result[sheet_name] = (records, data_type)
            logger.info("Excel sheet '%s': %d records, type=%s", sheet_name, len(records), data_type)
        return result
    except ImportError:
        # Fallback: try with xlrd for .xls
        logger.warning("openpyxl not available, attempting xlrd fallback")
        raise ValueError("openpyxl required for Excel parsing: pip install openpyxl")


def parse_json(content: str) -> Tuple[List[Dict], str]:
    """Parse JSON or JSONL content."""
    content = content.strip()
    if content.startswith("["):
        records = json.loads(content)
    elif content.startswith("{"):
        # Single record or JSONL
        try:
            records = [json.loads(content)]
        except json.JSONDecodeError:
            # Try JSONL
            records = [json.loads(line) for line in content.splitlines() if line.strip()]
    else:
        # JSONL
        records = [json.loads(line) for line in content.splitlines() if line.strip()]

    mapped_records: List[Dict] = []
    for idx, rec in enumerate(records, start=1):
        if not isinstance(rec, dict):
            continue
        mapped = _map_columns(rec)
        mapped["_lineage"] = {
            "source_type": "json",
            "row_number": idx,
        }
        mapped_records.append(mapped)
    records = mapped_records
    data_type = _detect_data_type(list(records[0].keys()) if records else [])
    logger.info("JSON parsed: %d records, type=%s", len(records), data_type)
    return records, data_type


def parse_sap_txt(content: str) -> Tuple[List[Dict], str]:
    """
    Parse SAP exported TXT files (ALV grid export, SM20, SU53, etc.)
    SAP exports are typically tab-delimited with separator lines (-----).
    """
    lines = content.splitlines()
    # Remove SAP header lines (lines starting with | or containing only dashes)
    data_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped or re.match(r"^[-|=+]+$", stripped):
            continue
        if stripped.startswith("|"):
            # Pipe-delimited table format
            stripped = stripped.strip("|")
        data_lines.append(stripped)

    if not data_lines:
        return [], "unknown"

    # Detect delimiter
    delimiter = _detect_delimiter("\n".join(data_lines[:10]))
    import csv
    reader = csv.DictReader(io.StringIO("\n".join(data_lines)), delimiter=delimiter)
    records = []
    for idx, row in enumerate(reader, start=2):
        mapped = _map_columns(row)
        mapped["_lineage"] = {
            "source_type": "sap_txt",
            "row_number": idx,
        }
        records.append(mapped)
    data_type = _detect_data_type(reader.fieldnames or [])
    logger.info("SAP TXT parsed: %d records, type=%s", len(records), data_type)
    return records, data_type


import hashlib

CONTROL_SCHEMAS = {
    'users': ['user_id', 'status', 'roles'],
    'changes': ['ticket_id', 'change_type', 'initiator', 'approver'],
    'transactions': ['invoice_id', 'vendor', 'invoice_amount'],
    'backup': ['job_name', 'status'],
    'incident': ['priority', 'status'],
    'interfaces': ['source_count', 'target_count']
}

class DataValidator:
    @staticmethod
    def validate(records: List[Dict], control_type: str) -> Dict[str, Any]:
        results = {}
        if not records:
            return {'pass': False, 'reason': 'No records provided'}

        # 1. Schema check
        required = set(CONTROL_SCHEMAS.get(control_type, []))
        if not required:
            results['pass'] = True
            return results

        columns = set(records[0].keys())
        missing = required - columns
        results['schema_valid'] = len(missing) == 0
        results['missing_columns'] = list(missing)
        
        if missing:
            results['pass'] = False
            results['reason'] = f'Missing required columns: {missing}'
            return results

        total_records = len(records)
        
        # 2. Completeness
        for col in required:
            null_count = sum(1 for r in records if not r.get(col) or str(r.get(col)).strip() == '')
            null_pct = null_count / total_records
            if null_pct > 0.005:
                results[f'{col}_null_pct'] = null_pct
                results['pass'] = False
                results['reason'] = f'Too many nulls in column {col} ({null_pct*100:.2f}%)'
                return results

        # 3. Row count and Hash
        results['row_count'] = total_records
        try:
            results['file_hash'] = hashlib.sha256(json.dumps(records, sort_keys=True).encode()).hexdigest()
        except:
            pass

        # 4. Duplicate check
        seen = set()
        duplicates = 0
        for r in records:
            row_str = str(sorted(r.items()))
            if row_str in seen:
                duplicates += 1
            else:
                seen.add(row_str)
        results['duplicates'] = duplicates

        results['pass'] = True
        return results

def parse_file(
    filename: str,
    content: Optional[str] = None,
    file_bytes: Optional[bytes] = None,
) -> Dict[str, Any]:
    """
    Auto-detect file format and parse.

    Returns:
        {
            "format": str,
            "sheets": {sheet_name: {"records": [...], "data_type": str}},
            "total_records": int,
            "primary_data_type": str,
        }
    """
    fname = filename.lower().strip()
    result = {"format": "unknown", "sheets": {}, "total_records": 0, "primary_data_type": "unknown"}

    try:
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            result["format"] = "excel"
            if file_bytes:
                sheets = parse_excel(file_bytes)
                for sheet_name, (records, data_type) in sheets.items():
                    val_result = DataValidator.validate(records, data_type)
                    result["sheets"][sheet_name] = {"records": records, "data_type": data_type, "validation": val_result}
                    result["total_records"] += len(records)
                if sheets:
                    result["primary_data_type"] = list(sheets.values())[0][1]

        elif fname.endswith(".json") or fname.endswith(".jsonl"):
            result["format"] = "json"
            if content:
                records, data_type = parse_json(content)
                val_result = DataValidator.validate(records, data_type)
                result["sheets"]["default"] = {"records": records, "data_type": data_type, "validation": val_result}
                result["total_records"] = len(records)
                result["primary_data_type"] = data_type

        elif fname.endswith(".txt") or "sap" in fname:
            result["format"] = "sap_txt"
            if content:
                records, data_type = parse_sap_txt(content)
                val_result = DataValidator.validate(records, data_type)
                result["sheets"]["default"] = {"records": records, "data_type": data_type, "validation": val_result}
                result["total_records"] = len(records)
                result["primary_data_type"] = data_type

        else:
            # Default: CSV
            result["format"] = "csv"
            if content:
                records, data_type = parse_csv(content)
                val_result = DataValidator.validate(records, data_type)
                result["sheets"]["default"] = {"records": records, "data_type": data_type, "validation": val_result}
                result["total_records"] = len(records)
                result["primary_data_type"] = data_type

    except Exception as e:
        logger.error("Parser error for %s: %s", filename, str(e))
        result["error"] = str(e)

    return result


def extract_user_roles(records: List[Dict]) -> Dict[str, List[str]]:
    """
    Convert flat user records to {user_id: [roles]} dict for SoD engine.
    Handles both wide format (one row per user) and long format (one row per role).
    """
    user_roles: Dict[str, List[str]] = {}

    for rec in records:
        uid = str(
            rec.get("user_id") or rec.get("username") or
            rec.get("user") or rec.get("logon") or "UNKNOWN"
        ).strip()
        if not uid or uid == "UNKNOWN":
            continue

        role = str(rec.get("roles") or rec.get("role") or rec.get("profile") or "").strip()
        if not role:
            continue

        if uid not in user_roles:
            user_roles[uid] = []

        # Handle comma/semicolon-separated roles in one cell
        for r in re.split(r"[,;|]", role):
            r = r.strip()
            if r and r not in user_roles[uid]:
                user_roles[uid].append(r)

    return user_roles
